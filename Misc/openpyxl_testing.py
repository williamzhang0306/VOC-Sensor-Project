#lets pull some stuff from that big csv
import openpyxl
from openpyxl import Workbook
from pathlib import Path

location = "/Users/williamzhang/Documents/MyPyStuff/Gas Sensor/VOCs signal vs time data.xlsx"
wb = openpyxl.load_workbook(location)

print(wb)


##tutorial

#create a workbook - one worksheet create
wb2 = Workbook()

#select the active work sheet
ws = wb2.active
ws1 = wb2.create_sheet("test_sheet")

ws.title = "New Title"

ws3 = wb2["New Title"]

print(wb2.sheetnames)

ws['A4'] = 4

ws.cell(row = 3, column = 1, value = False)

wb2.save('test_file.xlsx')

l=[]
x = 2
for y in range(1,10):
    v = ws.cell(row = x, column = y)
    l.append(v)

print(l)
